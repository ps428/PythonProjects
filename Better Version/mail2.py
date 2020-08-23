import email.message
import smtplib

#Opening excel
from openpyxl import load_workbook
fl=input("Enter filename: ")
if len(fl)<1:
	fl="mails.xlsx"
workbook = load_workbook(filename=fl)
##

msg = email.message.Message()

##SUBJECT HERE
msg['Subject'] = 'SUBJECT'

sender = 'ABC@gmail.com'
passw = 'PASSWORD'
msg['From'] = sender

TEXT = """
<html>
<body>

<p style="color: black;">Dear Sir,</p>

<p style="color: red;">Body</p>

<p style="color: black;">Thanking You,<br>

</body>
</html>

	"""

s= smtplib.SMTP(host='smtp.gmail.com', port=z587)
s.starttls()
s.login(sender, passw)

i=0
counter=input("Enter no. of recipeints: ")
for i in range(1,int(counter)+1): 
	#sheet = workbook.get_sheet_by_name('sheet')
	sheet = workbook.active

	rcv_email = sheet.cell(row=i,column=1)


	rcv_email=rcv_email.value

	print("Sending mail number {} to:".format(i),rcv_email)

	msg.add_header('Content-Type','text/html')
	msg.set_payload(TEXT)

	# Send the message via local SMTP server.
	#s = smtplib.SMTP('localhost')
	s.sendmail(msg['From'], rcv_email, msg.as_string())

s.quit()