import os, time
from pytube import YouTube
from openpyxl import load_workbook

fl=input("Enter filename: ")
if len(fl)<1:
	fl="MySirG.xlsx"
workbook = load_workbook(filename=fl)
##

counter = input("Enter no. of recipeints: ")
#start = input("Enter starting index: ")

for i in range(1,int(counter)+1): 
	#sheet = workbook.get_sheet_by_name('sheet')
	sheet = workbook.active

	address = sheet.cell(row=i,column=1)
	yt = YouTube(address.value)
	#os.system("youtube-dl -i "+address.value)
	#os.system("k")