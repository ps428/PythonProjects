# A Python program to send bulk mails by Pranav Soni.
# CAUTION: DO NOT RENAME THIS PYTHON FILE AS 'email.py', ANYTHING ELSE IS JUST FINE.
# CAUTION: DO NOT USE .value() OR .value() OR .values at line 46. Only use .value for this as it is not a method.
# Keep a file names 'emails.py' in same directory.
# Else you can also enter the excel filename which has mails.
# Daily mail limit is 500 mails. 
# It takes about 3-10 seconds to send a mail.
# Make sure mails are in Column A only, doesn't matter if they are hyperlinks. Else edits can be made at line 37.
# Enter sender's email id and password in 'sender_email' {line 26} and 'password' {line 27} variables. 
# Enter your message from line 52 and Subject at line 50.

import smtplib, ssl


#Opening excel
from openpyxl import load_workbook
fl=input("Enter filename: ")
if len(fl)<1:
	fl="excelFileName.xlsx"
workbook = load_workbook(filename=fl)
##


port = 587  # For starttls
smtp_server = "smtp.gmail.com"
sender_email = "ENTER SENDER'S EMAIL ADDERSS HERE"
password = "ENTER SENDER'S PASSWORD HERE"

##
#sheet = workbook.active
i=0
counter=input("Enter no. of recipeints: ")
for i in range(1,int(counter)+1): 
	#sheet = workbook.get_sheet_by_name('sheet')
	sheet = workbook.active

	rcv_email = sheet.cell(row=i,column=1)

# for value in sheet.iter_rows(min_row=1,
#                              max_row=1,
#                              min_col=1,
#                              max_col=1,	
#                              values_only=True):
# 	for en in value:
			
	receiver_email = rcv_email.value
	print("Sending mail number {} to:".format(i),receiver_email)
	#

	SUBJECT='SUBJECT GOES HERE'
	TEXT = """
THIS IS MAIL BODY

THANKS
	"""
	#
	message='Subject:{}\n\n{}'.format(SUBJECT,TEXT)


	context = ssl.create_default_context()
	with smtplib.SMTP(smtp_server, port) as server:
		server.ehlo()  # Can be omitted
		server.starttls(context=context)
		server.ehlo()  # Can be omitted
		server.login(sender_email, password)
		server.sendmail(sender_email, receiver_email, message)
	#

	


print("All mails sent")
workbook.save('Mailing List.xlsx')   
