# A Python program to send bulk emails with attachments by Pranav Soni.
# CAUTION: DO NOT RENAME THIS PYTHON FILE AS 'email.py', ANYTHING ELSE IS JUST FINE.
# CAUTION: DO NOT USE .value() OR .value() OR .values at line 45. Only use .value for this as it is not a method.
# Python code to illustrate Sending bulk mails with attachments. 
# Keep the attachment in same directory. Enter the name of attachment at line 70. If you don't want to keep file at same directory, then you may edit the directory location at line 72.
# Daily mail limit is roughly 500 mails. 
# Keep a file names 'emails.py' in same directory.
# Else you can also enter the excel filename which has mails.
# Make sure mails are in Column A only, doesn't matter if they are hyperlinks. Else edits can be made at line 44.
# It takes about 10-20 seconds to send a mail.# Enter Mail id and Password at lines 30 and 31 respectively.
#Enter mail Subject and Body at lines 57 and 59 respectively.

# libraries to be imported 
import smtplib 
import os

from email.mime.multipart import MIMEMultipart 
from email.mime.text import MIMEText 
from email.mime.base import MIMEBase 
from email import encoders 

##
from openpyxl import load_workbook
fl=input("Enter excel filename with extension or press ENTER for 'mails.xlsx': ")
if len(fl)<1:
	fl="mails.xlsx"
workbook = load_workbook(filename=fl)
##

fromaddr = "ENTER SENDER'S EMAIL HERE"#"EMAIL address of the sender"
password = "ENTER SENDER'S PASSWORD HERE"#"Password_of_the_sender"

# creates SMTP session 
s = smtplib.SMTP('smtp.gmail.com', 587) 
# start TLS for security 
s.starttls() 
# Authentication 
s.login(fromaddr, password) 

counter=input("Enter no. of recipeints: ")
for i in range(1,int(counter)+1): 
	#sheet = workbook.get_sheet_by_name('sheet')
	sheet = workbook.active
	x = sheet.cell(row=i,column=1) #Mail from sheet
	toaddr = x.value #CAUTION DON'T USE .value() ONLY USE .value

	print("Sending mail number {} to:".format(i), toaddr)

	# instance of MIMEMultipart 
	msg = MIMEMultipart() 

	# storing the senders email address 
	msg['From'] = fromaddr 
	# storing the receivers email address 
	msg['To'] = toaddr 
	# storing the subject 
	msg['Subject'] = "THIS IS SUBJECT"
	# string to store the body of the mail 
	body = """
BODY GOES HERE

Regards

	"""
	#"Body_of_the_mail"

	# attach the body with the msg instance 
	msg.attach(MIMEText(body, 'plain')) 
#ATTACHMENT	# open the file to be sent 
	filename = 'pythonlearn.pdf' #"File_name_with_extension"
	fileLocation = os.getcwd() + '/' + filename
	attachment = open(fileLocation)
	#attachment = open("/home/pranav/Desktop/Python Projects/Email Sender/Attachments/{}".format(filename), "rb") 
	print("File attached succesfully")
	# instance of MIMEBase and named as p 
	p = MIMEBase('application', 'octet-stream') 

	# To change the payload into encoded form 
	p.set_payload((attachment).read()) 

	# encode into base64 
	encoders.encode_base64(p) 

	p.add_header('Content-Disposition', "attachment; filename= %s" % filename) 
	# attach the instance 'p' to instance 'msg' 
	msg.attach(p) 
	
	# Converts the Multipart msg into a string 
	text = msg.as_string() 

	# sending the mail 
	s.sendmail(fromaddr, toaddr, text) 	
	#Editing Column B for as sent
	sentornot = sheet.cell(row=i,column=2)
	sentornot.value = 'Sent'
	

# terminating the session 
s.quit() 
workbook.save('Mailing List.xlsx')   
print("All Sent")